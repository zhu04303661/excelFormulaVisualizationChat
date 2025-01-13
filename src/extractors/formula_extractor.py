"""Excel公式提取器，用于提取和分析公式"""

import re
from openpyxl.utils import get_column_letter, column_index_from_string
from ..utils.cell_utils import is_yellow_cell, is_blue_cell, get_cell_address
from .header_extractor import HeaderExtractor
import ipdb;
from openpyxl import load_workbook
import networkx as nx
import matplotlib.pyplot as plt

import formulas
#from formulas.parser.nodes import Formula, Cell

class Node:
    _node_counter = 0  # 类变量，用于跟踪节点数量
    
    def __init__(self, cell_name, original_formula, cell_variable_name, variable_expression):
        Node._node_counter += 1
        self.index = Node._node_counter  # 为每个节点分配唯一的索引
        self.cell_name = cell_name
        self.original_formula = original_formula 
        self.cell_variable_name = cell_variable_name
        self.variable_expression = variable_expression
        self.children = []

    def __str__(self):
        return f"Node({self.index}: cell={self.cell_name}, vname={self.cell_variable_name}, formula={self.original_formula}, expr={self.variable_expression})"

    @classmethod
    def reset_counter(cls):
        """重置节点计数器"""
        cls._node_counter = 0

class FormulaExtractor:
    def __init__(self, workbook, excel_path=None):
        """
        初始化公式提取器
        
        Args:
            workbook: openpyxl工作簿对象
            excel_path: Excel文件路径，用于读取实际值
        """
        self.workbook = workbook
        self.excel_path = excel_path
        self.header_cache = {}
        self.input_cells = []
        self.output_cells = []  # 添加输出单元格列表
        self._init_header_cache()
        self.scan_input_cells()
        self.scan_output_cells()  # 添加扫描输出单元格的调用
        
    def _init_header_cache(self):
        """初始化标题缓存"""
        wb_data = load_workbook(self.excel_path, data_only=True)

        
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            header_extractor = HeaderExtractor(ws)
            sheet_cache = {}
            #ipdb.set_trace()
            # 扫描工作表中的所有单元格
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    row_header, col_header = header_extractor.find_nearest_header(row, col, 'both')
                    cell_key = f"{get_column_letter(col)}{row}"
                    
                    # 创建组合标题
                    combined_header = None
                    if row_header and col_header:
                        combined_header = f"{row_header}.{col_header}"
                    elif row_header:
                        combined_header = row_header
                    elif col_header:
                        combined_header = col_header
                        
                        
                    worksheet = wb_data[sheet_name]
                    actual_value = worksheet.cell(row, col).value
   
   
                    sheet_cache[cell_key] = {
                        'row_header': row_header,
                        'col_header': col_header,
                        'combined_header': combined_header,
                        'actual_value' : actual_value
                    }
                    
            wb_data.close()
            self.header_cache[sheet_name] = sheet_cache
            
    def _get_cell_cached_headers(self, cell):
        sheet_name = cell.parent.title
        cell_address = cell.coordinate
        return self._get_cached_headers(sheet_name, cell_address)
    
    def _get_cached_headers(self, sheet_name, cell_address):
        """
        从缓存中获取单元格的标题
        
        Args:
            sheet_name: 工作表名称
            cell_address: 单元格地址（如 'A1'）
            
        Returns:
            tuple: (行标题, 列标题, 组合标题)
        """
        if sheet_name in self.header_cache and cell_address in self.header_cache[sheet_name]:
            headers = self.header_cache[sheet_name][cell_address]
            return headers['row_header'], headers['col_header'], headers['combined_header'],headers['actual_value']
        else:
            print('没有找到缓存中的标题',sheet_name,cell_address)
            return None, None, None
    
    def extract_formulas(self):
        """提取所有工作表中的公式"""
        formulas = []
        
        # 创建一个启用了data_only的工作簿副本来获取计算后的值
        #wb_data = load_workbook(self.excel_path, data_only=True)
        #ipdb.set_trace()
        for sheet_name in self.workbook.sheetnames:
            ws = self.workbook[sheet_name]
            header_extractor = HeaderExtractor(ws)
            #data_worksheet = wb_data[sheet_name]

            # 遍历所有单元格
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    formula = cell.value
                    
                    # 检查是否是公式且单元格不是黄色背景
                    if (isinstance(formula, str) and formula.startswith('=') and 
                        not is_yellow_cell(cell)):
                        # 获取列字母和单元格地址
                        col_letter = get_column_letter(col)
                        cell_address = f'{col_letter}{row}'
                        
                        # 查找最近的行标题和列标题
                        row_header, col_header, combined_header,actual_value = self._get_cached_headers(sheet_name, cell_address)
                        
                        # 如果缓存中没有找到，则实时查找
                        if row_header is None:
                            row_header, col_header = header_extractor.find_nearest_header(row, col, 'both')
                            combined_header = f"{row_header}.{col_header}" if row_header and col_header else (row_header or col_header)
                        
                        #if cell_address in ['AE6','AE7']:
                        #    ipdb.set_trace()

                        # 获取当前表格名称（蓝色背景的单元格）
                        table_name = self._find_table_name(cell, ws, row)
            
                                              
                        #print('\n==========开始分解公式...单元格 ',cell)
                        #if cell_address=='AE6' or cell_address=='AE6':
                        #    ipdb.set_trace()
                        # 分解公式
                        decomposed_formula = self._decompose_formula(ws, formula)
                        
                        #print('\n==========开始转换...单元格 ',cell)
                        
                        # 将分解后的公式转换为变量表达式
                        variable_expr = self._convert_to_variable_expression(  ws, f"={decomposed_formula}" )
                        
                        #print(table_name,'公式',cell_address,' 为：', combined_header,variable_expr)
                        # 创建公式信息字典
                        formula_info = {
                            '工作表': sheet_name,
                            '表格名称': table_name or '未分类',
                            '单元格': cell_address,
                            '行列组合标题': combined_header or '无标题',
                            '原计算方法': '公式'+formula,
                            '变量表达式': combined_header+''+variable_expr,
                            '输出变量': combined_header,
                            '计算方法': variable_expr,
                            '当前值': actual_value  # 添加实际值

                        }
                        
                        formulas.append(formula_info)
        # 关闭data_only工作簿
        #wb_data.close()
        return formulas
    
    def _find_table_name(self, cell, worksheet, current_row):
        """
        查找当前行所属的表格名称（向上查找直到遇到合并单元格）
        
        Args:
            cell: 目标单元格对象
            worksheet: 工作表对象
            current_row: 当前行号
            
        Returns:
            str: 表格名称，如果未找到则返回None
        """
        # 获取所有合并单元格范围
        merged_ranges = worksheet.merged_cells.ranges
        
        # 获取目标单元格所在列
        target_col = cell.column
        
        # 向上查找直到遇到合并单元格
        for row in range(current_row-1, 0, -1):
            check_cell = worksheet.cell(row=row, column=target_col)
            
            # 检查当前单元格是否在任何合并单元格范围内
            for merged_range in merged_ranges:
                if check_cell.coordinate in merged_range:
                    # 获取合并单元格的起始单元格（左上角）的值
                    start_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
                    return start_cell.value if start_cell.value else None
        
        return None
    
    def _analyze_formula_dependencies(self,cells):
        """
        分析公式依赖关系并生成最终的计算表达式
        
        Args:
            cells (list): 需要分析的单元格列表
            formula_extractor: FormulaExtractor实例
        
        Returns:
            list: 包含依赖关系的公式列表
        """
        print("\n正在分析公式依赖关系...")
        
        # 创建依赖图
        dependency_graph = {}
        final_formulas = []

        for item in cells:
            cell = item['cell']
            formula = cell.value  # 获取单元格的公式
            worksheet = cell.parent  # 获取单元格所属的工作表
            
            # 获取公式依赖
            basic_cells, new_formula,tree,path = self._trace_formula_dependencies(
                worksheet, 
                formula,
                cell,
                cell.row,
                cell.column
            )
            # 将分解后的公式转换为变量表达式
            variable_new_formula = self._convert_to_variable_expression(
                worksheet, f"={new_formula}"
            )

            
            # 创建公式信息字典
            formula_info = {
                '工作表': worksheet.title,
                '单元格': cell.coordinate,
                '原始公式': formula,
                '标题组合': item.get('header', ''),  # 使用item中的header信息
                '合并公式': new_formula,
                '变量公式': variable_new_formula,
                '基础单元格': basic_cells,
                '路径': path,
                '依赖树': tree
           #     '依赖变量': dependencies['依赖变量'],
    #            '计算顺序': dependencies['计算顺序'],
    #            '完整表达式': dependencies['完整表达式']
            }
            
            final_formulas.append(formula_info)
            
            # # 更新依赖图
            # dependency_graph[formula_info['标题组合']] = {
            #     'deps': dependencies['依赖变量'],
            #     'expr': dependencies['完整表达式']
            # }
        
        # 按照依赖关系排序
        #sorted_formulas = sort_by_dependencies(final_formulas)
        
        return final_formulas

    def _trace_formula_dependencies(self, worksheet, formula, cell, row, col):
        """
        追踪公式依赖关系
        
        Args:
            worksheet: 工作表对象
            formula: 公式字符串
            cell: 当前单元格
            row: 当前行号
            col: 当前列号
            
        Returns:
            tuple: (基础单元格集合, 简化表达式, 依赖树, 节点路径)
        """
        basic_cells = set()
        path = []  # 现在存储Node对象
        nodelist = []
        current_node = Node(
            cell_name=self.get_cell_coordinate_with_sheet(cell),
            original_formula=cell.value,
            cell_variable_name= self._get_cell_cached_headers(cell),
            variable_expression=self._convert_to_variable_expression(worksheet, f"{cell.value}"),
        )
        to_process = [(worksheet, formula, cell, row, col,current_node)]

        visited = set()
        simplified_formula = formula[1:] if formula.startswith('=') else formula
        new_formula = ''

        while to_process:
            current_ws, current_formula, current_cell, current_row, current_col,current_node = to_process.pop(0)
            
            current_cell_name = self.get_cell_coordinate_with_sheet(current_cell)
            
            if current_cell_name in visited:
                continue

            visited.add(current_cell_name)
            print(' 当前处理的单元格： ', current_cell_name, '=', current_formula)

            # 创建当前节点
            variable_expr = self._convert_to_variable_expression(current_ws, f"{current_formula}")

            path.append(current_node.__str__())
            nodelist.append(current_node)
            
            # if cell.coordinate == 'E37':
            #     ipdb.set_trace()
            
            # 分解公式
            new_current_formula = self._decompose_formula(current_ws, current_formula)

            #print('分解后的公式', decomposed_formula)
            
            #ipdb.set_trace()      
            

            # 提取单元格引用并重组公式
            new_formula = self._replace_cell_refs(new_formula, current_cell, current_ws)
       
            refs = self._extract_cell_refs(new_current_formula, current_ws)
  
            print(' 待合成次数： ', len(to_process))
            if len(to_process) > 3000:
                ipdb.set_trace()
                break

            #ipdb.set_trace()      
            for ref in refs:
                self._process_cell_reference(current_ws, ref, basic_cells, to_process, current_node.children)
        #ipdb.set_trace()    
        if len(nodelist) > 0:
            print(f"准备绘制公式树，单元格: {cell.coordinate}")
            print(f"节点列表长度: {len(nodelist)}")
            print(f"根节点信息: {nodelist[0]}")
            self.visualize_interactive_formula_tree(nodelist[0], f'formula_tree_{cell.coordinate}.html')
        else:
            print(f"警告：单元格 {cell.coordinate} 没有生成节点列表")
        #print(cell.coordinate, ' 合成后的公式 ', new_formula)
        return basic_cells, new_formula, nodelist, path
    

    
    
    def _extract_cell_refs(self, formula,worksheet):
        """提取公式中的单元格引用"""
        # 匹配模式：
        # 1. 可选的+号
        # 2. 工作表名（可能包含中文）
        # 3. 感叹号
        # 4. 单元格引用
        variable_expr = formula[1:] if formula.startswith('=') else formula
        
        # 修改正则表达式以匹配单元格引用，并确保所有引用都带有工作表名
        cell_refs = re.findall(
            r"('?[^+\-*/(),\s!]+?'?!\$?[A-Z]+\$?\d+|\$?[A-Z]+\$?\d+)", 
            variable_expr
        )
        
        # 处理提取到的单元格引用，确保都带有工作表名
        processed_refs = []
        seen_refs = set()  # 用于跟踪已处理的引用
        current_sheet = None
        
        for ref in cell_refs:
            processed_ref = None
            if '!' in ref:
                # 如果包含工作表名，保存当前工作表名
                current_sheet = ref.split('!')[0].strip("'")
                processed_ref = ref
            else:
                # 如果不包含工作表名，使用最近的工作表名或当前工作表
                sheet_prefix = f"{current_sheet or worksheet.title}!"
                # 移除可能存在的$符号
                clean_ref = ref.replace('$', '')
                processed_ref = f"{sheet_prefix}{clean_ref}"
            
            # 只有当引用未被处理过时才添加
            if processed_ref not in seen_refs:
                seen_refs.add(processed_ref)
                processed_refs.append(processed_ref)
        
        # 用处理后的引用替换原来的cell_refs
        cell_refs = processed_refs
        
        return cell_refs
    
    def _get_cell_range_references(self, worksheet, start_ref, end_ref):
        """
        获取单元格范围内的所有单元格引用
        
        Args:
            worksheet: 工作表对象
            start_ref: 起始单元格引用
            end_ref: 结束单元格引用
            
        Returns:
            list: 单元格引用列表
        """
        start_col = ''.join(c for c in start_ref if c.isalpha())
        start_row = int(''.join(c for c in start_ref if c.isdigit()))
        end_col = ''.join(c for c in end_ref if c.isalpha())
        end_row = int(''.join(c for c in end_ref if c.isdigit()))
        
        # 确保行列范围正确
        if start_row > end_row:
            start_row, end_row = end_row, start_row
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        if start_col_idx > end_col_idx:
            start_col_idx, end_col_idx = end_col_idx, start_col_idx
        
        # 生成单元格引用列表
        cell_refs = []
        for r in range(start_row, end_row + 1):
            for c in range(start_col_idx, end_col_idx + 1):
                cell_refs.append(f"{worksheet.title}!{get_column_letter(c)}{r}")
        
        return cell_refs
    
    def _process_cell_reference(self, worksheet, ref, basic_cells, to_process, children):
        """
        处理单个单元格引用
        
        Args:
            worksheet: 工作表对象
            ref: 单元格引用
            basic_cells: 基础单元格集合
            to_process: 待处理的单元格列表
            children: 当前节点的子节点列表
        """
        try:
            if '!' in ref:
                sheet_name, cell_ref = ref.split('!')
                sheet_name = sheet_name.strip("'")
                sheet_name = sheet_name.strip("=")
                if sheet_name not in self.workbook.sheetnames:
                    children.append(Node(
                        cell_name=ref,
                        original_formula=f"跳过无效工作表: {ref}",
                        cell_variable_name= "ERROR",
                        variable_expression="ERROR"
                    ))
                    return
                target_ws = self.workbook[sheet_name]
            else:
                cell_ref = ref
                target_ws = worksheet
            
            col_str = ''.join(filter(str.isalpha, cell_ref))
            row_num = int(''.join(filter(str.isdigit, cell_ref)))
            ref_col = column_index_from_string(col_str)
            
            if not (1 <= row_num <= target_ws.max_row and 
                   1 <= ref_col <= target_ws.max_column):
                node = Node(
                    cell_name=ref,
                    original_formula=f"跳过无效单元格引用: {ref}",
                    cell_variable_name= "ERROR",
                    variable_expression="ERROR"
                )
                children.append(node)
                return
            
            cell = target_ws.cell(row=row_num, column=ref_col)
            full_ref = f"{target_ws.title}!{cell_ref}"
            
            
            if is_yellow_cell(cell):
                basic_cells.add(full_ref)
                #ipdb.set_trace()
                node = Node(
                    cell_name=self.get_cell_coordinate_with_sheet(cell),
                    original_formula=cell.value,
                    cell_variable_name= self._get_cell_cached_headers(cell),
                    variable_expression="INPUT"  # 或其他标识输入单元格的值
                )
                children.append(node)
            elif isinstance(cell.value, str) and cell.value.startswith('='):
                node = Node(
                    cell_name=self.get_cell_coordinate_with_sheet(cell),
                    original_formula=str(cell.value),
                    cell_variable_name= self._get_cell_cached_headers(cell),
                    variable_expression=str(cell.value)
                )
                children.append(node)
                to_process.append((target_ws, cell.value, cell, row_num, ref_col,node))
            else:
                basic_cells.add(full_ref)

            
        except Exception as e:
            node = Node(
                cell_name=ref,
                original_formula=f"处理单元格引用出错 {ref}: {str(e)}",
                cell_variable_name= "ERROR",
                variable_expression="ERROR"
            )
            children.append(node)
    
    def _analyze_formula(self, formula):
        """
        分析公式的计算方法
        
        Args:
            formula: 公式字符串
            
        Returns:
            str: 公式的数学表达式
        """
        # 去掉公式开头的等号
        formula = formula[1:] if formula.startswith('=') else formula
        
        # 基本运算符及其数学符号
        operators = {
            '+': '+',
            '-': '−',  # 使用数学减号
            '*': '×',
            '/': '÷',
            '^': '⁰¹²³⁴⁵⁶⁷⁸⁹',  # 上标数字
            '=': '=',
            '>': '>',
            '<': '<',
            '>=': '≥',
            '<=': '≤',
            '<>': '≠'
        }
        
        # 替换运算符
        for op, math_op in operators.items():
            formula = formula.replace(op, f' {math_op} ')
        
        # 处理函数
        formula = re.sub(r'SUM\((.*?)\)', r'∑(\1)', formula, flags=re.IGNORECASE)
        formula = re.sub(r'AVERAGE\((.*?)\)', r'⟨\1⟩', formula, flags=re.IGNORECASE)
        formula = re.sub(r'PRODUCT\((.*?)\)', r'∏(\1)', formula, flags=re.IGNORECASE)
        formula = re.sub(r'SQRT\((.*?)\)', r'√(\1)', formula, flags=re.IGNORECASE)
        formula = re.sub(r'ABS\((.*?)\)', r'|\1|', formula, flags=re.IGNORECASE)
        
        return formula.strip() 
    
    def _convert_to_variable_expression(self, worksheet, formula):
        """
        将公式转换为使用行标题作为变量的表达式
        """
        variable_expr = formula[1:] if formula.startswith('=') else formula
        
        # 修改正则表达式以匹配单元格引用，并确保所有引用都带有工作表名
        cell_refs = re.findall(
            r"('?[^+\-*/(),\s!]+?'?!\$?[A-Z]+\$?\d+|\$?[A-Z]+\$?\d+)", 
            variable_expr
        )
        
        # 处理提取到的单元格引用，确保都带有工作表名
        processed_refs = []
        seen_refs = set()  # 用于跟踪已处理的引用
        current_sheet = None
        
        for ref in cell_refs:
            processed_ref = None
            if '!' in ref:
                # 如果包含工作表名，保存当前工作表名
                current_sheet = ref.split('!')[0].strip("'")
                processed_ref = ref
            else:
                # 如果不包含工作表名，使用最近的工作表名或当前工作表
                sheet_prefix = f"{current_sheet or worksheet.title}!"
                # 移除可能存在的$符号
                clean_ref = ref.replace('$', '')
                processed_ref = f"{sheet_prefix}{clean_ref}"
            
            # 只有当引用未被处理过时才添加
            if processed_ref not in seen_refs:
                seen_refs.add(processed_ref)
                processed_refs.append(processed_ref)
        
        # 用处理后的引用替换原来的cell_refs
        cell_refs = processed_refs
        
        # 创建替换映射
        replacements = {}
        for cell_ref in sorted(cell_refs, key=len, reverse=True):  # 按长度降序排序
            target_ws = worksheet
            cell_address = cell_ref
            sheet_name = worksheet.title
            
            if '!' in cell_ref:
                parts = cell_ref.split('!')
                sheet_name = parts[0].strip("'")
                cell_address = parts[1]
                if sheet_name in self.workbook.sheetnames:
                    target_ws = self.workbook[sheet_name]
            
            # 从缓存中获取标题
            row_header, col_header, combined_header,actual_value = self._get_cached_headers(sheet_name, cell_address)
            if combined_header:
                replacements[cell_ref] = combined_header
            elif row_header:  # 如果没有组合标题，则使用行标题作为后备
                replacements[cell_ref] = row_header
        
        # 替换单元格引用为变量名
        result_parts = []
        current_pos = 0
        
        # 保持运算符和括号
        for match in re.finditer(r'([+\-*/(),]|\s+|[^+\-*/(),\s]+)', variable_expr):
            token = match.group(0)
            if '$' in token:
                # 如果不包含工作表名，使用最近的工作表名或当前工作表
                sheet_prefix = f"{current_sheet or worksheet.title}!"
                # 移除可能存在的$符号
                clean_ref = ref.replace('$', '')
                token = f"{sheet_prefix}{clean_ref}"
    
            
            if token.strip() in replacements:
                result_parts.append(replacements[token.strip()])
            else:
                # 保留运算符和括号
                if token in ['+', '-', '*', '/', '(', ')', ',']:
                    result_parts.append(f' {token} ')
                elif not token.isspace():  # 跳过空白字符
                    result_parts.append(token)
        
        # 组合结果
        variable_expr = ''.join(result_parts)
        
        # 清理多余的空格
        variable_expr = re.sub(r'\s+', ' ', variable_expr).strip()
        
        # 美化最终表达式
        variable_expr = f"= {variable_expr}"
        
        return variable_expr
    
    def _decompose_formula(self, worksheet, formula):
        """
        分解公式，将SUM、AVERAGE、IRR、NPV等函数展开为基本运算
        
        Args:
            worksheet: 当前工作表对象（公式所在的工作表）
            formula: 公式字符串
        """
        if not formula.startswith('='):
            return formula
            
        formula = formula[1:]  # 移除等号
        
        #if 'IF' in formula:
        #    ipdb.set_trace()
        
        # 依次处理各种函数
        formula = self._decompose_irr(worksheet, formula)
        formula = self._decompose_npv(worksheet, formula)
        formula = self._decompose_sum(worksheet, formula)
        formula = self._decompose_average(worksheet, formula)
        formula = self._add_missing_sheet_references(worksheet, formula)
        
        return formula

    def _decompose_irr(self, worksheet, formula):
        """处理IRR函数"""
        irr_pattern = r'IRR\((.*?)\)'
        processed_positions = set()  # 记录已处理的位置
        
        while 'IRR' in formula.upper():
            match = re.search(irr_pattern, formula, re.IGNORECASE)
            if not match or match.start() in processed_positions:
                break
            
            range_str = match.group(1)
            if ':' in range_str:
                cells = self._expand_range_reference(worksheet, range_str)
                replacement = f"IRR({','.join(cells)})"
                start_pos = match.start()
                formula = formula[:start_pos] + replacement + formula[match.end():]
                processed_positions.add(start_pos)  # 记录已处理的位置
            else:
                processed_positions.add(match.start())  # 标记已检查但无需处理的位置
            
        return formula

    def _decompose_npv(self, worksheet, formula):
        """处理NPV函数"""
        npv_pattern = r'NPV\((.*?),(.*?)\)'
        processed_positions = set()  # 记录已处理的位置
        
        while 'NPV' in formula.upper():
            match = re.search(npv_pattern, formula, re.IGNORECASE)
            if not match or match.start() in processed_positions:
                break
            
            rate = match.group(1)
            range_str = match.group(2)
            
            if ':' in range_str:
                cells = self._expand_range_reference(worksheet, range_str)
                replacement = f"NPV({rate},{','.join(cells)})"
                start_pos = match.start()
                formula = formula[:start_pos] + replacement + formula[match.end():]
                processed_positions.add(start_pos)  # 记录已处理的位置
            else:
                processed_positions.add(match.start())  # 标记已检查但无需处理的位置
            
        return formula

    def _decompose_sum(self, worksheet, formula):
        """处理SUM函数"""
        sum_pattern = r'SUM\((.*?)\)'
        while 'SUM' in formula.upper():
            match = re.search(sum_pattern, formula, re.IGNORECASE)
            if not match:
                break
            
            range_str = match.group(1)
            if ':' in range_str:
                cells = self._expand_range_reference(worksheet, range_str)
                replacement = f"({' + '.join(cells)})"
                formula = formula.replace(match.group(0), replacement)
            else:
                # 处理单个单元格或逗号分隔的单元格
                cells = self._expand_comma_separated_refs(worksheet, range_str)
                replacement = f"({' + '.join(cells)})"
                formula = formula.replace(match.group(0), replacement)
            
        return formula

    def _decompose_average(self, worksheet, formula):
        """处理AVERAGE函数"""
        avg_pattern = r'AVERAGE\((.*?)\)'
        while 'AVERAGE' in formula.upper():
            match = re.search(avg_pattern, formula, re.IGNORECASE)
            if not match:
                break
            
            range_str = match.group(1)
            if ':' in range_str:
                cells = self._expand_range_reference(worksheet, range_str)
                cell_count = len(cells)
                replacement = f"(({' + '.join(cells)})/{cell_count})"
                formula = formula.replace(match.group(0), replacement)
            else:
                # 处理单个单元格或逗号分隔的单元格
                cells = self._expand_comma_separated_refs(worksheet, range_str)
                cell_count = len(cells)
                replacement = f"(({' + '.join(cells)})/{cell_count})"
                formula = formula.replace(match.group(0), replacement)
            
        return formula

    def _expand_range_reference(self, worksheet, range_str):
        """展开单元格范围引用为单个单元格列表"""
        if '!' in range_str:
            sheet_name, cell_range = range_str.split('!')
            sheet_name = sheet_name.strip("'")
            target_ws = self.workbook[sheet_name]
            start_ref, end_ref = cell_range.split(':')
        else:
            target_ws = worksheet
            start_ref, end_ref = range_str.split(':')
        
        return self._get_cell_range_references(target_ws, start_ref, end_ref)

    def _expand_comma_separated_refs(self, worksheet, refs_str):
        """展开逗号分隔的单元格引用"""
        cells = []
        for cell_ref in refs_str.split(','):
            cell_ref = cell_ref.strip()
            if '!' in cell_ref:
                # 如果包含工作表引用，保持原样
                cells.append(cell_ref)
            else:
                # 如果没有工作表引用，添加当前工作表名称
                cells.append(f"{worksheet.title}!{cell_ref}")
        return cells

    def _add_missing_sheet_references(self, worksheet, formula):
        """为缺少工作表引用的单元格添加引用"""
        cell_pattern = r'([A-Z]+[0-9]+)'
        
        def replace_cell_ref(match):
            cell_ref = match.group(1)
            # 检查前面是否已经有工作表引用
            start = match.start(1)
            if start > 0 and formula[start-1] == '!':
                # 如果前面有!，说明已经有工作表引用，直接返回原引用
                return cell_ref
            return f"{worksheet.title}!{cell_ref}"
        
        return re.sub(cell_pattern, replace_cell_ref, formula)

    def scan_input_cells(self):
        """
        扫描所有工作表中的黄色背景单元格（输入单元格）
        并将结果保存到文本文件
        
        Returns:
            list: 输入单元格信息列表
        """
        # 使用传入的文件路径创建data_only工作簿
        #wb_data = load_workbook(self.excel_path, data_only=True)
        input_cells = []
        
        # 扫描所有工作表
        for sheet_name in self.workbook.sheetnames:
            worksheet = self.workbook[sheet_name]
            # 获取对应的data_only工作表
            #data_worksheet = wb_data[sheet_name]
            print(f'\n正在扫描工作表输入单元格: {sheet_name}')
            
            # 扫描所有单元格
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # 检查是否是黄色背景的单元格
                    if is_yellow_cell(cell):
                        # 从data_only工作表获取实际值
                        #ipdb.set_trace()
                        #data_cell = data_worksheet.cell(row=row, column=col)
                        #actual_value = data_cell.value
                        
                        # 获取表格名称
                        table_name = self._find_table_name(cell, worksheet, row)
                        
                        # 获取标题组合
                        # 查找最近的行标题和列标题
                        row_header, col_header, combined_header,actual_value = self._get_cached_headers(sheet_name, cell_address=cell.coordinate)
                        
                        # 记录输入单元格信息
                        input_cells.append({
                            '工作表': sheet_name,
                            '单元格': cell.coordinate,
                            '表格名称': table_name,
                            '标题组合': combined_header,
                            '当前值': actual_value
                        })
        
        # 关闭data_only工作簿
        #wb_data.close()
        
        # 按表格名称和单元格位置排序
        sorted_input_cells = sorted(input_cells, 
                                  key=lambda x: (x['表格名称'] or '', x['工作表'], x['单元格']))
        #ipdb.set_trace()
        self.input_cells = sorted_input_cells
        return 

    def scan_output_cells(self):
        """
        扫描所有工作表中的绿色背景单元格（输出单元格）
        并将结果保存到文本文件
        
        Returns:
            list: 输出单元格信息列表
        """
        # 使用传入的文件路径创建data_only工作簿
        #wb_data = load_workbook(self.excel_path, data_only=True)
        output_cells = []
        
        # 扫描所有工作表
        for sheet_name in self.workbook.sheetnames:
            if sheet_name != '测算结果输出':
                continue
            
            worksheet = self.workbook[sheet_name]
            # 获取对应的data_only工作表
            #data_worksheet = wb_data[sheet_name]
            print(f'\n正在扫描工作表输出单元格: {sheet_name}')
            
            # 扫描所有单元格
            for row in range(1, worksheet.max_row + 1):
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row, column=col)
                    formula = cell.value
                    
                    # 检查是否是公式且单元格不是黄色背景
                    if (isinstance(formula, str) and formula.startswith('=') and not is_yellow_cell(cell)):                        # 从data_only工作表获取实际值
                        #data_cell = data_worksheet.cell(row=row, column=col)
                        #actual_value = data_cell.value
                        
                        # 获取表格名称
                        table_name = self._find_table_name(cell, worksheet, row)
                        
                        # 获取标题组合
                        row_header, col_header, combined_header,actual_value = self._get_cached_headers(sheet_name, cell_address=cell.coordinate)
                        
                        # 记录输出单元格信息
                        output_cells.append({
                            'cell': cell,
                            '工作表': sheet_name,
                            '单元格': cell.coordinate,
                            '表格名称': table_name,
                            '标题组合': combined_header,
                            '当前值': actual_value
                        })
        
        # 关闭data_only工作簿
        #wb_data.close()
        
        # 按表格名称和单元格位置排序
        self.output_cells = sorted(output_cells, 
                                 key=lambda x: (x['表格名称'] or '', x['工作表'], x['单元格']))
        
        return self.output_cells

    def _reassemble_formula(self, original_formula, cell_refs, worksheet):
        """
        重组Excel公式，将引用的单元格替换为其对应的公式
        
        Args:
            original_formula: 原始公式字符串
            cell_refs: 提取的单元格引用列表
            worksheet: 当前工作表对象
        
        Returns:
            str: 重组后的公式
        """
        new_formula = original_formula
        
        for ref in cell_refs:
            # 解析工作表名和单元格地址
            if '!' in ref:
                sheet_name, cell_addr = ref.split('!')
                sheet_name = sheet_name.strip("'")
                target_ws = self.workbook[sheet_name]
            else:
                target_ws = worksheet
                cell_addr = ref
            
            # 获取目标单元格
            col_letter = ''.join(filter(str.isalpha, cell_addr))
            row_num = int(''.join(filter(str.isdigit, cell_addr)))
            target_cell = target_ws.cell(row=row_num, column=column_index_from_string(col_letter))
            
            # 如果目标单元格包含公式，递归处理
            if isinstance(target_cell.value, str) and target_cell.value.startswith('='):
                sub_formula = self._decompose_formula(target_ws, target_cell.value)
                new_formula = new_formula.replace(ref, f"({sub_formula})")
            else:
                # 如果是普通值，直接替换
                cell_value = target_cell.value
                if isinstance(cell_value, (int, float)):
                    new_formula = new_formula.replace(ref, str(cell_value))
                else:
                    new_formula = new_formula.replace(ref, f"'{str(cell_value)}'")
        
        return new_formula

    def _replace_original_formula_part_by_cell_decode_formula(self, original_formula, target_part, new_formula_part, worksheet):
        """
        使用formula_part替换original_formula中对应的单元格引用，并用括号包裹替换的部分
        
        Args:
            original_formula (str): 原始公式
            target_part (str): 需要被替换的单元格引用
            new_formula_part (str): 用于替换的公式部分
            worksheet: 当前工作表对象
            
        Returns:
            str: 替换后的公式
        """
        new_formula = original_formula  
        print(' 替换前的公式： ', new_formula)
        print(' 要替换的部分： ', target_part)
        
        # 如果new_formula_part不是简单的数字，则用括号包裹
        if new_formula_part and not new_formula_part.strip('-+').replace('.', '').isdigit():
            new_formula_part = f"({new_formula_part})"
            
        # 执行替换
        new_formula = new_formula.replace(target_part, new_formula_part)
                
        print(' 替换后的公式： ', new_formula)
        return new_formula

    def _replace_cell_refs(self, original_formula, cell, worksheet):
        """
        使用子单元格的公式替换原公式中的单元格引用
        
        Args:
            original_formula: 原始公式字符串
            cell: 当前单元格
            cell_refs: 提取的单元格引用列表
            worksheet: 当前工作表对象
            
        Returns:
            str: 替换后的公式
        """
        if original_formula == '':
            original_formula = self.get_cell_coordinate_with_sheet(cell)
            
        new_formula = original_formula

        #for ref in cell_refs:
        ref = self.get_cell_coordinate_with_sheet(cell)            # 解析工作表名和单元格地址
        


        if '!' in ref:
            sheet_name, cell_addr = ref.split('!')
            sheet_name = sheet_name.strip("'")
            target_ws = self.workbook[sheet_name]
        else:
            target_ws = worksheet
            cell_addr = ref.lstrip('+')  # 移除可能的前导加号
        
        # 获取目标单元格
        col_letter = ''.join(filter(str.isalpha, cell_addr))
        row_num = int(''.join(filter(str.isdigit, cell_addr)))
        #if ref in original_formula:
        #    target_cell = target_ws.cell(row=row_num, column=column_index_from_string(col_letter))
        #else:
        target_cell = cell
        #    ref = self.get_cell_coordinate_with_sheet(cell)

        
        # 获取单元格的公式或值
        if isinstance(target_cell.value, str) and target_cell.value.startswith('='):
            # 递归处理公式
            sub_formula = self._decompose_formula(target_ws, target_cell.value)
            # 保持原有的运算符（如果存在）
            operator = '+' if ref.startswith('+') else ''
            #replacement = f"{operator}({cell.value[1:]})"  # 移除子公式的等号
            #formula = str(target_cell.value)
            replacement = f"{operator}({sub_formula})"  # 移除子公式的等号
        else:
            # 如果是普通值，保持原有格式
            cell_value = target_cell.value
            if isinstance(cell_value, (int, float)):
                replacement = ref  # 保持原有引用
            else:
                replacement = f"'{str(cell_value)}'"
        
        # 替换公式中的引用
        new_formula = new_formula.replace(ref, replacement)
        # if cell.coordinate == 'E14':
        #     print(' 替换前的公式： ', original_formula)
        #     print(' 要替换的部分： ', ref)
        #     print(' 替换后的公式： ', new_formula)
        #     ipdb.set_trace()
        
        return new_formula

    def get_cell_coordinate_with_sheet(self, cell):
        return f'{cell.parent.title}!{cell.coordinate}'

    def _expand_formula(self, workbook, current_sheet_name, cell_address, visited=None):
        if visited is None:
            visited = set()

        sheet = workbook[current_sheet_name]
        cell = sheet[cell_address]

        # 防止循环引用
        if (current_sheet_name, cell.coordinate) in visited:
            return cell.coordinate  # 或者返回一些默认值，避免无限递归
        visited.add((current_sheet_name, cell.coordinate))

        if cell.data_type == 'f':  # 判断是否为公式
            formula = cell.value
            # 移除等号
            if formula.startswith('='):
                formula = formula[1:]

            # 正则表达式匹配单元格引用，包括跨工作表引用
            # 支持的格式：
            # - Sheet1!A1
            # - 'My Sheet'!$A$1
            # - A1
            # - $A$1

            # 在公式中，可能存在函数名称或文本，需要排除掉
            pattern = r"('([^']|'')+'|[A-Za-z0-9_]+)?(?:!)?(\$?[A-Za-z]+\$?\d+)"
            matches = re.findall(pattern, formula)

            for full_match, sheet_name_group, cell_ref in matches:
                # 确定工作表名称
                if '!' in full_match:
                    if sheet_name_group.startswith("'") and sheet_name_group.endswith("'"):
                        ref_sheet_name = sheet_name_group[1:-1].replace("''", "'")
                    else:
                        ref_sheet_name = sheet_name_group
                else:
                    # 如果未指定工作表名称，则默认使用当前工作表
                    ref_sheet_name = current_sheet_name

                ref_cell_address = cell_ref

                # 检查工作表是否存在
                if ref_sheet_name in workbook.sheetnames:
                    ref_sheet = workbook[ref_sheet_name]
                else:
                    continue  # 或者处理不存在的工作表的情况

                # 获取引用的单元格
                ref_cell = ref_sheet[ref_cell_address]
                if ref_cell.data_type == 'f' or ref_cell.data_type == 'n' or ref_cell.data_type == 's':
                    expanded = self._expand_formula(workbook, ref_sheet_name, ref_cell_address, visited)
                    # 用括号包裹，以保持运算优先级
                    formula = formula.replace(f"{full_match}{'!' if full_match else ''}{cell_ref}", f'({expanded})', 1)
                else:
                    # 如果无法识别单元格类型，保留原始引用
                    pass

            return formula
        else:
            # 如果单元格不是公式，返回其值
            if cell.value is None:
                return ''
            else:
                return str(cell.value)

    #def visualize_formula_tree(self, root_node, output_path='formula_tree.png'):
        """
        将公式依赖树可视化为图片,使用分层布局从上到下展示
        """
        try:
            # 设置中文字体
            plt.rcParams['font.sans-serif'] = ['SimHei','Songti SC','STFangsong'] # 用来正常显示中文标签
            plt.rcParams['axes.unicode_minus'] = False  # 用来正常显示负号
            
            # 创建新图形
            plt.figure(figsize=(20, 15))  # 加大图形尺寸以适应层级布局
            
            # 创建有向图
            G = nx.DiGraph()
            
            def add_nodes_edges(node, parent_id=None):
                """递归添加节点和边"""
                if node is None:
                    return
                
                # 创建节点标签
                node_label = f"{node.cell_name}\n{node.cell_variable_name[2]}\n{node.variable_expression}"
                
                # 添加节点
                G.add_node(node.index, label=node_label)
                
                # 如果有父节点,添加边
                if parent_id is not None:
                    G.add_edge(parent_id, node.index)
                
                # 递归处理所有子节点
                for child in node.children:
                    add_nodes_edges(child, node.index)
            
            # 从根节点开始递归构建整个树
            add_nodes_edges(root_node)
            
            # 使用分层布局,从上到下排列
            pos = nx.spring_layout(G, k=2)  # 首先用spring布局获取初始位置
            
            # 获取每个节点的层级
            levels = nx.shortest_path_length(G, root_node.index)
            
            # 调整y坐标使节点按层级从上到下排列
            y_max = max(levels.values())
            for node in pos:
                pos[node] = (pos[node][0], 1 - (levels[node] / y_max))
            
            # 绘制节点和边
            nx.draw(G, pos,
                   node_color='lightblue',
                   node_size=5000,  # 增大节点尺寸
                   arrows=True,
                   edge_color='gray',
                   width=2,
                   with_labels=False,
                   arrowsize=20)  # 增大箭头尺寸
                   
            # 添加节点标签
            labels = nx.get_node_attributes(G, 'label')
            nx.draw_networkx_labels(G, pos, labels, font_size=8)
            
            # 调整布局,增加上下边距
            plt.margins(y=0.2)
            
            # 保存图片
            plt.savefig(output_path, bbox_inches='tight', dpi=300)
            
            # 显示图形
            plt.show()
            
        except Exception as e:
            print(f"绘制公式树时出错: {str(e)}")
            import traceback
            traceback.print_exc()

    def visualize_interactive_formula_tree(self, root_node, output_path):
        """
        将公式依赖树可视化为交互式HTML页面
        
        Args:
            root_node: 树的根节点
            output_path: HTML文件保存路径
        """
        try:
            def node_to_dict(node):
                """将Node对象转换为字典格式"""
                brief_info, detail_info = self._format_node_info(node)
                
                return {
                    'id': str(node.index),
                    'brief': brief_info,
                    'detail': detail_info,
                    'children': [node_to_dict(child) for child in node.children]
                }
            
            # 将树结构转换为JSON格式
            tree_data = node_to_dict(root_node)
            
            # 读取HTML模板
            import os
            template_path = os.path.join(os.path.dirname(__file__), 'show.html')
            with open(template_path, 'r', encoding='utf-8') as f:
                html_content = f.read()
            
            # 将树数据注入到HTML中
            import json
            tree_data_json = json.dumps(tree_data, ensure_ascii=False)
            #ipdb.set_trace()
            # 整行替换
            html_content = html_content.replace(
                'const sampleData = {};',  # 匹配整行
                f'const sampleData = {tree_data_json};'
            )          
              
            # 保存生成的HTML文件
            with open(output_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            print(f"公式树已保存到: {output_path}")
            
            # 自动在浏览器中打开生成的HTML文件
            import webbrowser
            webbrowser.open('file://' + os.path.abspath(output_path))
            
        except Exception as e:
            print(f"生成HTML公式树时出错: {str(e)}")
            import traceback
            traceback.print_exc()

    def _format_node_info(self, node):
        """
        格式化节点信息，返回简略和详细两种格式
        
        Args:
            node: 节点对象
            
        Returns:
            tuple: (brief_info, detail_info)
                brief_info: 简略信息，包含单元格地址、公式类型和当前值
                detail_info: 详细信息，包含完整的公式内容
        """
        try:
            # 提取公式类型（如SUM, AVERAGE等）
            formula_type = ''
            if node.original_formula:
                if isinstance(node.original_formula, (int, float)):
                    formula_type = '[数值]'
                elif isinstance(node.original_formula, str) and node.original_formula.startswith('='):
                    # 提取第一个函数名
                    #match = re.match(r'=([A-Z]+)', node.original_formula)
                    match = re.match(r"^=([A-Z]+)$$.*$$", node.original_formula)

                    if match:
                        # 复杂函数，只显示函数名
                        formula_type = f"[{match.group(1)}]"
                    else:
                        # 简单四则运算，显示完整公式
                        # 去掉等号，保留运算部分
                        simple_formula = node.original_formula
                        # 如果公式不太长，直接显示
                        if len(simple_formula) <= 30:  # 可以调整长度阈值
                            formula_type = f"[{simple_formula}]"
                        else:
                            # 如果太长，截断显示
                            formula_type = f"[{simple_formula[:27]}...]"
            
            # 格式化当前值
            current_value = node.cell_variable_name[3] if node.cell_variable_name[3] is not None else 'N/A'
            
            # 构建简略信息
            brief_info = f" {node.cell_variable_name[2]} {formula_type} = {current_value}"
            
            # 构建详细信息
            detail_info = f"【概要】{brief_info} \n【详细】 {node.cell_name}: {node.original_formula}  {node.cell_variable_name[2]}  当前值:{node.cell_variable_name[3]}",
            
            return brief_info, detail_info
            
        except Exception as e:
            print(f"格式化节点信息时出错: {str(e)}")
            return f"{node.cell_name}", {'cell': node.cell_name, 'error': str(e)}

    