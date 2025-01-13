"""Excel公式分析主程序"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from src.extractors.formula_extractor import FormulaExtractor
from src.analyzers.formula_analyzer import FormulaAnalyzer
import traceback
#import ipdb
import os
import argparse
import traceback
from openpyxl import load_workbook
from src.extractors.formula_extractor import FormulaExtractor


def sort_formulas(formulas):
    """
    对公式列表进行排序：
    1. 首先按照表格名称分组
    2. 然后按照变量表达式中等号后面的变量个数排序
    
    Args:
        formulas (list): 公式信息列表
    
    Returns:
        list: 排序后的公式列表
    """
    def count_variables(expr):
        """计算变量表达式中等号后面的变量个数"""
        if not expr or '=' not in expr:
            return 0
        # 获取等号后面的部分
        right_side = expr.split('=', 1)[1]
        # 计算点号的出现次数（每个变量都包含一个点号）
        return right_side.count('.')
    
    def get_sort_key(formula):
        """获取排序键"""
        table_name = formula.get('表格名称', '')
        var_count = count_variables(formula.get('变量表达式', ''))
        return (table_name, var_count)
    
    # 对公式列表进行排序
    sorted_formulas = sorted(formulas, key=get_sort_key)
    return sorted_formulas

def print_and_save_formulas(formulas):
    """
    打印公式分析结果并保存到文本文件
    
    Args:
        formulas (list): 排序后的公式列表
    """
    output_text = []
    current_table = None
    
    print("\n排序后的公式列表：")
    output_text.append("排序后的公式列表：\n")
    
    for formula in formulas:
        # 当遇到新的表格名称时，打印分隔行
        if formula['表格名称'] != current_table:
            current_table = formula['表格名称']
            separator = f"\n{'='*80}\n表格名称: {current_table}\n{'='*80}"
            print(separator)
            output_text.append(separator)

        formula_info = f"""
{formula['变量表达式']}  
"""
        
        print(formula_info)
        output_text.append(formula_info)
    
    # 保存到文本文件
    output_file = 'formula_analysis.txt'
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_text))
    
    print(f"\n分析结果已保存到: {output_file}")

def save_input_cells_to_text(input_cells, output_file='input_cells.txt'):
    """
    将输入单元格信息保存到文本文件
    
    Args:
        input_cells (list): 输入单元格信息列表
        output_file (str): 输出文件路径，默认为'input_cells.txt'
    """
    if not input_cells:
        print("没有找到输入单元格")
        return
        
    output_text = []
    current_table = None
    
    output_text.append("输入单元格列表：\n")
    #ipdb.set_trace()
    for cell_info in input_cells:
        # 当遇到新的表格名称时，添加分隔行
        if cell_info['表格名称'] != current_table:
            current_table = cell_info['表格名称']
            separator = f"\n{'='*80}\n表格名称: {current_table}\n{'='*80}"
            output_text.append(separator)
        
        # 格式化输出每个输入单元格的信息
        cell_info_text = f"""
输入变量名称: {cell_info['标题组合']}
说明：已输入值: {cell_info['当前值']}  位置: {cell_info['工作表']}!{cell_info['单元格']}
"""
        #     cell_info_text = f"""{cell_info['标题组合']}"""
        output_text.append(cell_info_text)
    
    # 保存到文本文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_text))
    
    print(f"\n找到 {len(input_cells)} 个输入单元格")
    print(f"输入单元格信息已保存到: {output_file}")

def save_output_cells_to_text(output_cells, output_file='output_cells.txt'):
    """
    将输出单元格信息保存到文本文件
    
    Args:
        output_cells (list): 输出单元格信息列表
        output_file (str): 输出文件路径，默认为'output_cells.txt'
    """
    if not output_cells:
        print("没有找到输出单元格")
        return
        
    output_text = []
    current_table = None
    
    output_text.append("输出单元格列表：\n")
    
    for cell_info in output_cells:
        # 当遇到新的表格名称时，添加分隔行
        if cell_info['表格名称'] != current_table:
            current_table = cell_info['表格名称']
            separator = f"\n{'='*80}\n表格名称: {current_table}\n{'='*80}"
            output_text.append(separator)
        
        # 格式化输出每个输出单元格的信息
        cell_info_text = f"""
输出变量名称: {cell_info['标题组合']}
说明：计算结果: {cell_info['当前值']}  位置: {cell_info['工作表']}!{cell_info['单元格']}
{'-'*80}"""
        output_text.append(cell_info_text)
    
    # 保存到文本文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_text))
    
    print(f"\n找到 {len(output_cells)} 个输出单元格")
    print(f"输出单元格信息已保存到: {output_file}")

def save_output_formulas_to_text(output_formulas, output_file='output_formulas.txt'):
    """
    将输出单元格信息保存到文本文件
    
    Args:
        output_formulas (list): 输出单元格信息列表
        output_file (str): 输出文件路径，默认为'output_formulas.txt'
    """
    if not output_formulas:
        print("没有找到输出")
        return
        
    output_text = []
    current_table = None
    
    output_text.append("输出单元格列表：\n")
    
    
    for formula_info in output_formulas:
                  
    #         # 创建公式信息字典
    #        # 创建公式信息字典
    #         formula_info = {
    #             '工作表': worksheet.title,
    #             '单元格': cell.coordinate,
    #             '原始公式': formula,
    #             '标题组合': item.get('header', ''),  # 使用item中的header信息
    #             '合并公式': new_formula,
    #             '变量公式': variable_new_formula,
    #             '基础单元格': basic_cells,
    #             '路径': path,
    #             '依赖树': tree
    #        #     '依赖变量': dependencies['依赖变量'],
    # #            '计算顺序': dependencies['计算顺序'],
    # #            '完整表达式': dependencies['完整表达式']
    #         }

        
        # 格式化输出每个输出单元格的信息
        cell_info_text = f"""
输出变量名称: {formula_info['标题组合']}
说明：计算公式: {formula_info['变量公式']}  
路径: {formula_info['路径']}
依赖树: {formula_info['依赖树']}
{'-'*80}"""
        output_text.append(cell_info_text)
    
    # 保存到文本文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_text))
    
    print(f"\n找到 {len(output_formulas)} 个输出目标公式")
    print(f"输出单元格信息已保存到: {output_file}")


def save_results_to_excel(formulas, output_file):
    """
    将分析结果保存到Excel文件
    
    Args:
        formulas (list): 公式信息列表
        output_file (str): 输出文件路径
    """
    # 创建DataFrame
    df = pd.DataFrame(formulas)
    
    # 根据实际数据动态获取列名
    if formulas:
        columns = list(formulas[0].keys())
        # 确保某些重要列排在前面（如果存在的话）
        priority_columns = ['层级', '工作表', '表格名称', '单元格', '行标题', '列标题']
        ordered_columns = (
            [col for col in priority_columns if col in columns] +
            [col for col in columns if col not in priority_columns]
        )
        df = df[ordered_columns]
    
    # 保存到Excel
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='公式分析结果')
        
        # 调整列宽
        worksheet = writer.sheets['公式分析结果']
        for idx, col in enumerate(df.columns, 1):
            max_length = max(
                df[col].astype(str).apply(len).max(),
                len(str(col))
            )
            worksheet.column_dimensions[get_column_letter(idx)].width = max_length + 2

def save_synthesized_formulas(analyzer, formulas, input_cells, output_cells, output_file='synthesized_formulas.txt'):
    """
    保存合成后的公式到文本文件
    
    Args:
        analyzer: FormulaAnalyzer实例
        formulas: 所有公式的列表
        input_cells: 输入单元格列表
        output_cells: 输出单元格列表
        output_file: 输出文件路径
    """
    output_text = ["合成公式列表：\n"]
    current_table = None
    
    for output_cell in output_cells:
        # 当遇到新的表格名称时，添加分隔行
        if output_cell['表格名称'] != current_table:
            current_table = output_cell['表格名称']
            separator = f"\n{'='*80}\n表格名称: {current_table}\n{'='*80}"
            output_text.append(separator)
        
        # 获取输出单元格的完整引用
        cell_ref = f"{output_cell['工作表']}!{output_cell['单元格']}"
        
        # 合成公式
        synthesized = analyzer.synthesize_formula(cell_ref, formulas, input_cells)
        
        # 格式化输出
        formula_text = f"""
输出变量: {output_cell['标题组合']}
位置: {cell_ref}
合成公式: {synthesized}
{'-'*80}"""
        output_text.append(formula_text)
    
    # 保存到文本文件
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write('\n'.join(output_text))
    
    print(f"\n合成公式已保存到: {output_file}")



def sort_by_dependencies(formulas):
    """
    根据依赖关系对公式进行排序
    
    Args:
        formulas (list): 包含依赖信息的公式列表
    
    Returns:
        list: 排序后的公式列表
    """
    # 创建依赖图
    graph = {}
    for formula in formulas:
        if formula['标题组合']:  # 只处理有标题组合的公式
            graph[formula['标题组合']] = set(formula['依赖变量'])
    
    # 拓扑排序
    result = []
    visited = set()
    temp_visited = set()
    
    def visit(node):
        if node in temp_visited:
            raise ValueError(f"检测到循环依赖: {node}")
        if node in visited:
            return
        
        temp_visited.add(node)
        for dep in graph.get(node, set()):
            if dep in graph:  # 只访问存在的节点
                visit(dep)
        temp_visited.remove(node)
        visited.add(node)
        result.append(node)
    
    # 对每个节点进行深度优先搜索
    for formula in formulas:
        if formula['标题组合'] and formula['标题组合'] in graph:
            try:
                visit(formula['标题组合'])
            except ValueError as e:
                print(f"警告: {e}")
    
    # 根据拓扑排序重新组织公式列表
    sorted_formulas = []
    for node in result:
        for formula in formulas:
            if formula['标题组合'] == node:
                sorted_formulas.append(formula)
                break
    
    # 添加未参与排序的公式（没有标题组合的）
    for formula in formulas:
        if not formula['标题组合'] or formula['标题组合'] not in graph:
            sorted_formulas.append(formula)
    
    return sorted_formulas

def process_excel_formulas(input_file, output_file):
    """
    处理Excel文件中的公式
    
    Args:
        input_file (str): 输入Excel文件路径
        output_file (str): 输出Excel文件路径
    """
    try:
        # 加载工作簿
        print('正在加载Excel文件...')
        workbook = load_workbook(input_file, data_only=False)
        
        # 提取公式
        print('正在提取公式...')
        formula_extractor = FormulaExtractor(workbook, input_file)
        #formulas = formula_extractor.extract_formulas()
        
        #if formulas:
        if formula_extractor:
#            print(f'\n找到 {len(formulas)} 个公式')
            
            # 对公式进行排序
            #print('\n正在对公式进行排序...')
            #sorted_formulas = sort_formulas(formulas)
            
            # 打印和保存分析结果到文本文件
            #print_and_save_formulas(sorted_formulas)
            
            # 保存输入单元格信息到文本文件
            save_input_cells_to_text(formula_extractor.input_cells)
            
            # 保存输出单元格信息到文本文件
            save_output_cells_to_text(formula_extractor.output_cells)
            
            
            output_formulas = formula_extractor._analyze_formula_dependencies(formula_extractor.output_cells)
            save_output_formulas_to_text(output_formulas)
            
            # 保存结果到Excel
            print('\n正在保存分析结果到Excel...')
            #save_results_to_excel(sorted_formulas, output_file)
            
            print(f'\nExcel结果已保存到: {output_file}')
            
        else:
            print('未找到任何公式')
            
    except Exception as e:
        print(f'处理过程出现错误: {str(e)}')
        print('\n详细错误信息:')
        print(traceback.format_exc())
    
            
def main():
    """
    主函数，处理命令行参数并执行公式处理
    """
    # 设置命令行参数ˆ
    parser = argparse.ArgumentParser(description='Excel公式分析工具')
    parser.add_argument('--input_file',  
                        default='input.xlsx',
                        help='输入Excel文件路径')
    parser.add_argument('--output', '-o', 
                       default='formula_analysis_result.xlsx',
                       help='输出Excel文件路径 (默认: formula_analysis_result.xlsx)')
    
    # 解析命令行参数
    args = parser.parse_args()
    

    # 检查输入文件是否存在
    if not os.path.exists(args.input_file):
        print(f'错误: 输入文件不存在: {args.input_file}')
        return
    
    try:
        # 处理Excel公式
        process_excel_formulas(args.input_file, args.output)
    except Exception as e:
        print(f'程序执行出错: {str(e)}')
        print('\n详细错误信息:')
        print(traceback.format_exc())

if __name__ == '__main__':
    main()